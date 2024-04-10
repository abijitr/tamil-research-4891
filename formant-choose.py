import openpyxl
import sys

def get_first_middle_last(array):
    try:
        # Calculate the middle index
        middle_index = len(array) // 2

        # If the array length is even, adjust the middle index to the left
        if len(array) % 2 == 0:
            middle_index -= 1

        # Grab the first, middle, and last items
        first_item = array[0]
        middle_item = array[middle_index]
        last_item = array[-1]

        return first_item, middle_item, last_item
    except IndexError:
        print("IndexError occurred. Array:", array)

english_words_set = set(['Butter', 'Valet', 'Tip', 'Puny', 'Tall', 'Tiny', 'Cup', 'Drag', 'Gala', 'Mouse'])
tamil_words_set = ([
    'Kannaa',
    'Pattam : Kite',
    'Vellai',
    'Kullam : pond',
    'Padattum',
    'Pallam : pit',
    'Pillai : Son',
    'Kannaadi : mirror',
    'Ketta : Bad',
    'Vettu : cut / chop',
    'Kollai',
    'Aattam : Dance',
    'Vattum : circle',
    'Mattum',
    'Sattam',
    'Vellam : Flood',
    'Vannam : color',
    'Kulaai : Pipe',
    'Ullai : inside',
    'Kattam : Crossword',
    'Ennai : oil',
    'Tunnivu : courage',
    'Anna : Elder Brother',
    'Kannam : cheek',
    'Thanni : water',
    'Katti'
])

total_words_set = english_words_set.union(tamil_words_set)

# Load the Excel workbook
workbook = openpyxl.load_workbook('split-formant_data.xlsx')

# Create a new workbook to store the split worksheets
new_workbook = openpyxl.Workbook()

for worksheet in workbook:
    print(f"Processing worksheet: {worksheet.title}")
    new_worksheet = new_workbook.create_sheet(title=worksheet.title)
    new_worksheet.append([None, "Time_s_Before", "F1_Hz_Before", "F2_Hz_Before", "F3_Hz_Before", "F4_Hz_Before", None, None, None, None, None, None, None, None, "Time_s_After", "F1_Hz_After", "F2_Hz_After", "F3_Hz_After", "F4_Hz_After"])
    # Specify the range of rows to iterate over
    start_row = 1
    end_row = worksheet.max_row  # Adjust as needed
    
    # Iterate through the rows using a counted for loop
    for row_number in range(start_row, end_row + 1):
        # Access individual row using the row number
        row = worksheet[row_number]
        # Check if any cell in the row contains a word from the sets
        contains_word = False
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if cell.value.strip() in total_words_set:
                    contains_word = True
                    break  # Exit the loop once a word is found
        if contains_word:
            new_worksheet.append([row[0].value])
            before_arr = []
            counter = 1
            ind_row = worksheet[row_number + counter]
            ind_row_values = [cell.value for cell in ind_row]
            while not all(element is None for element in ind_row_values): 
                before_arr.append(ind_row)
                ind_row = worksheet[row_number + counter]
                ind_row_values = [cell.value for cell in ind_row]
                counter+=1
            after_arr = []
            counter += 1
            ind_row = worksheet[row_number + counter]
            ind_row_values = [cell.value for cell in ind_row]
            while not all(element is None for element in ind_row_values):
                after_arr.append(ind_row)
                ind_row = worksheet[row_number + counter]
                ind_row_values = [cell.value for cell in ind_row]
                counter+=1
            
            if before_arr != [] and after_arr != []:
                first_before, second_before, third_before = get_first_middle_last(before_arr)
                first_after, second_after, third_after = get_first_middle_last(after_arr)
                first = first_before + first_after
                second = second_before + second_after
                third = third_before + third_after
                first = [cell.value for cell in first]
                second = [cell.value for cell in second]
                third = [cell.value for cell in third]
                # print(f"First row: {first}")
                # print(f"Second row: {second}")
                # print(f"Third row: {third}")
                new_worksheet.append(first)
                new_worksheet.append(second)
                new_worksheet.append(third)

new_workbook.save('summarized-formant-data.xlsx')
