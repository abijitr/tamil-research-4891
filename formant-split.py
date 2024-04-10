import openpyxl

# Load the Excel workbook
workbook = openpyxl.load_workbook('data-formant.xlsx')

# Specify the array of words
experiment_types = ['English 1', 'Tamil 1 102bpm', 'Tamil 1 60bpm', 'Tamil 2 60bpm', 'Tamil 2 102bpm', 'English 2']

# Create a new workbook to store the split worksheets
split_workbook = openpyxl.Workbook()

# Iterate through each original worksheet
for original_worksheet in workbook:
    # Create a dictionary to store worksheets for each word in the current worksheet
    word_worksheets = {}
    word_worksheet = None
    
    # Iterate through the rows of the current worksheet
    for row in original_worksheet.iter_rows(min_row=1, values_only=True):
        # Get the word from the first column
        word = row[0]

        # Check if the word matches any word in the array
        if word in experiment_types:
            # Check if the worksheet for the word exists, if not, create it
            if word not in word_worksheets:
                word_worksheets[word] = split_workbook.create_sheet(title=f"{original_worksheet.title}_{word}")

            # Get the worksheet for the word
            word_worksheet = word_worksheets[word]

            # Append the row to the worksheet
            word_worksheet.append(row)
        elif word_worksheet is not None:
            word_worksheet.append(row)

# Save the new workbook with split worksheets
split_workbook.save('split-formant_data.xlsx')

import subprocess

# Specify the path to the Python file you want to run
python_file_path = 'formant-choose.py'

# Run the Python file
subprocess.run(['python', python_file_path])